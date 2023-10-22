import datetime

import click
import xlrd
from openpyxl import load_workbook

from ynab.fineco import Transaction, AccountTransaction, AccountTransactionsOutput


@click.group()
@click.version_option()
def cli():
    "CLI tool to support data import and export from YNAB"


@cli.group()
def fineco():
    "Fineco related commands"


@fineco.command(name="describe-account-transactions")
@click.argument(
    "excel-file-name",
)
@click.option(
    "-o",
    "--output-format",
    help="Output format",
    type=click.Choice(["table", "csv"]),
    default="table",
)
def describe_account_transactions(excel_file_name: str, output_format: str):
    "Read an .xlsx file containing bank account transactions and output the in a table or a CSV file"
    click.echo("Here is some output")

    workbook = load_workbook(filename=excel_file_name)
    ws = workbook.active

    transactions = []
    for row in ws.iter_rows(min_row=8, max_col=7):
        t = AccountTransaction(
            date=datetime.datetime.strptime(row[0].value, '%d/%m/%Y').date(),
            amount=row[1].value or row[2].value,
            description=row[3].value,
            description_full=row[4].value,
            state=row[5].value,
            moneymap_category=row[6].value,
        )

        transactions.append(t)

    output = AccountTransactionsOutput(transactions)
    if output_format == "table":
        click.echo(output.table())
    elif output_format == "csv":
        click.echo(output.csv())


@fineco.command(name="read-credit-card")
@click.argument(
    "excel-file",
)
@click.option(
    "-o",
    "--option",
    help="An example option",
)
def read_credit_card(excel_file, option):
    "Read a credit card statement and output a table or a CSV file"
    click.echo("Here is some output")

    # Open the workbook
    workbook = xlrd.open_workbook(excel_file)

    # Select the first sheet (index 0) from the workbook
    sheet = workbook.sheet_by_index(0)

    transactions = []

    # Read data from cells with data
    for row in range(3, sheet.nrows):

        cell_value = sheet.cell_value(row, 1)
        if cell_value == "":
            continue
            
        owner = sheet.cell_value(row, 1)
        card_number = sheet.cell_value(row, 2)

        transaction_date = datetime.datetime(*xlrd.xldate_as_tuple(sheet.cell_value(row, 3), workbook.datemode))
        registration_date = datetime.datetime(*xlrd.xldate_as_tuple(sheet.cell_value(row, 4), workbook.datemode))

        description = sheet.cell_value(row, 5)
        operation_state = sheet.cell_value(row, 6)
        operation_type = sheet.cell_value(row, 7)
        circuit = sheet.cell_value(row, 8)
        transaction_type = sheet.cell_value(row, 9)
        amount = sheet.cell_value(row, 10)

        # print(f"{owner} {card_number} {transaction_date} {registration_date} {description} {operation_state} {circuit} {transaction_type} [{amount}]")
        # print(f"{owner} {card_number} {transaction_date} {registration_date} {description} {operation_type} {circuit} {transaction_type} [{amount}]")
        # print(f"{transaction_date} {description} [{amount}] ({circuit})")

        # for col in range(1, sheet.ncols):
        #     cell_value = sheet.cell_value(row, col)
        #         print(cell_value)

        transaction = Transaction(
            owner=owner,
            card_number=card_number,
            transaction_date=transaction_date,
            registration_date=registration_date,
            description=description,
            operation_state=operation_state,
            operation_type=operation_type,
            circuit=circuit,
            transaction_type=transaction_type,
            amount=amount,
        )

        transactions.append(transaction)


    for t in sorted(transactions, key=lambda t: t.amount):
        print(t.amount, t.description)